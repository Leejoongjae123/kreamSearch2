import openpyxl
import pandas as pd
import subprocess
import shutil
from bs4 import BeautifulSoup
import time
import datetime
import sys
import os
import math
import requests
import re
import random
import numpy
import datetime
import json
import pprint
from collections import defaultdict
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import smtplib  # SMTP 사용을 위한 모듈
from email.mime.multipart import MIMEMultipart  # 메일의 Data 영역의 메시지를 만드는 모듈
from email.mime.text import MIMEText  # 메일의 본문 내용을 만드는 모듈
from email.mime.base import MIMEBase
from email import encoders

# 특수 문자를 제외한 문자만 추출하는 함수
def extract_characters(text):
    return ''.join(re.findall(r'[\w\s]', text))

def find_value_by_key(d, target_key):
    if isinstance(d, dict):
        if target_key in d:
            return d[target_key]
        for key, value in d.items():
            if isinstance(value, dict):
                result = find_value_by_key(value, target_key)
                if result is not None:
                    return result
    elif isinstance(d, list):
        for item in d:
            result = find_value_by_key(item, target_key)
            if result is not None:
                return result
    return None

def find_values_by_key2(data, target_key):
    found_values = []

    def recurse(d):
        if isinstance(d, dict):
            for key, value in d.items():
                if key == target_key:
                    found_values.append(value)
                elif isinstance(value, dict):
                    recurse(value)
                elif isinstance(value, list):
                    for item in value:
                        if isinstance(item, dict):
                            recurse(item)

    recurse(data)
    return found_values

def remove_special_characters(input_string):
    # 정규 표현식을 사용하여 특수 문자를 제거합니다.
    clean_string = re.sub(r'[^a-zA-Z0-9가-힣\s]', '', input_string)
    return clean_string

def chrome_browser(url):
    chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]  # 크롬 버전을 확인한다.
    driver_path = r'.\\{chrome_ver}\\chromedriver.exe'

    if os.path.exists(driver_path):
        print(f"chromedriver is installed: {driver_path}")  # 있는 버전을 쓴다.
    else:
        print(f"install the chrome driver(ver: {chrome_ver})")  # 크롬을 최신 버전으로 설치한다.
        chromedriver_autoinstaller.install(True)

    options = webdriver.ChromeOptions()  # 크롬 옵션을 추가한다.
    # options.add_argument('headless')
    options.add_experimental_option("detach", True)  # 크롬 안 꺼지는 옵션 추가
    options.add_experimental_option("excludeSwitches", ["enable-logging"])  # 크롬 안 꺼지는 옵션 추가

    browser = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=options)  # 크롬 드라이버를 할당
    browser.get(url)
    browser.maximize_window()
    browser.implicitly_wait(3)
    return browser

def GetCGTransaction(daysAgo,token,refreshToken,headers,productNo):
    count=1
    dataList=[]
    cookies = {
        'i18n_redirected': 'kr',
        'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
        'AF_SYNC': '1701661188073',
        'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
        '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
        '_fbp': 'fb.2.1701936854500.2073430007',
        '_gid': 'GA1.3.1820798567.1701936855',
        'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
        'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
        'NA_SAS': '1',
        'NVADID': '0zC0001-KB5zvCorg1i4',
        'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
        'refresh_token_cookie': refreshToken,
        'csrf_refresh_token': '9b39c907-2d44-4947-952b-83e8b2b2cf55',
        'login_type': 'email',
        '_token.local': token,
        '_refresh_token.local': refreshToken,
        'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
        'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
        'strategy': 'local',
        'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
        'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'airbridge_utm_timestamp': '1701959996934',
        '_ga': 'GA1.3.248227678.1701661186',
        '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701962031505%2C%22c%22%3A1701959994791%2C%22l%22%3A1701960231505%7D',
        'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MDIzMTUwOSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODglN0Q=',
        'wcs_bt': 's_59a6a417df3:1701960232',
        '_gat_gtag_UA_153398119_1': '1',
        'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701960232744%7D',
        '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
        '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
    }
    while True:
        params = {
            'cursor': str(count),
            'per_page': '50',
            'sort': '',
            'request_key': 'b9576e81-7724-41dc-a75d-dedebc2bbe98',
        }
        try:
            response = requests.get(
                'https://kream.co.kr/api/p/products/{}/sales'.format(productNo),
                params=params,
                cookies=cookies,
                headers=headers,
            )
            print(response.text)
            results=json.loads(response.text)['items']
        except:
            print("더없다")
            break

        for result in results:
            try:
                price=result['price']
            except:
                price=""
            # print(price)
            try:
                size=extract_characters(result['option'])
            except:
                size=""
            # print(size)
            try:
                immediate=result['is_immediate_delivery_item']
            except:
                immediate=""
            # print(immediate)
            try:
                transactionDate=result['date_created']
                transactionTimestamp=datetime.datetime.strptime(transactionDate, "%Y-%m-%dT%H:%M:%SZ").timestamp()
            except:
                transactionDate=""
            # print(transactionDate)

            timeLimit=(datetime.datetime.now()-datetime.timedelta(days=daysAgo)).timestamp()
            if transactionTimestamp<timeLimit:
                print("날짜지남")
                break
            data={'category':"CG",'price':price,'size':size,'immediate':immediate,'transactionDate':transactionDate}
            # print(data)
            dataList.append(data)
        count+=1
        time.sleep(random.randint(10,15)*0.1)
        seen = set()
        for d in dataList:
            # 딕셔너리를 문자열로 변환
            dict_str = str(d)

            # 이미 본 딕셔너리인지 확인
            if dict_str in seen:
                # 중복된 딕셔너리가 발견되면 함수를 빠져나갑니다.
                return dataList

        if transactionTimestamp<timeLimit:
            break
        time.sleep(2)
    return dataList

def GetPMTransaction(PMScroll,token,refreshToken,headers,productNo):
    count=1
    dataList=[]

    cookies = {
        'i18n_redirected': 'kr',
        'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
        'AF_SYNC': '1701661188073',
        'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
        '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
        '_fbp': 'fb.2.1701936854500.2073430007',
        '_gid': 'GA1.3.1820798567.1701936855',
        'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
        'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
        'NA_SAS': '1',
        'NVADID': '0zC0001-KB5zvCorg1i4',
        'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
        'refresh_token_cookie': refreshToken,
        'csrf_refresh_token': '9b39c907-2d44-4947-952b-83e8b2b2cf55',
        'login_type': 'email',
        '_token.local': token,
        '_refresh_token.local': refreshToken,
        'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
        'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
        'strategy': 'local',
        'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
        'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'airbridge_utm_timestamp': '1701959996934',
        '_ga': 'GA1.3.248227678.1701661186',
        '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701962031505%2C%22c%22%3A1701959994791%2C%22l%22%3A1701960231505%7D',
        'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MDIzMTUwOSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODglN0Q=',
        'wcs_bt': 's_59a6a417df3:1701960232',
        '_gat_gtag_UA_153398119_1': '1',
        'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701960232744%7D',
        '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
        '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
    }

    while True:


        params = {
            'cursor': count,
            'per_page': '50',
            'sort': '',
            'request_key': 'b9576e81-7724-41dc-a75d-dedebc2bbe98',
        }
        try:
            response = requests.get('https://kream.co.kr/api/p/products/{}/asks'.format(productNo), params=params, cookies=cookies,
                                    headers=headers)
            print(response.text)
            results=json.loads(response.text)['items']
        except:
            print('더없다')
            break
        # pprint.pprint(results)
        for result in results:
            try:
                price=result['price']
            except:
                price=""
            # print(price)
            try:
                size=extract_characters(result['option'])
            except:
                size=""
            # print(size)
            try:
                quantity=result['quantity']
            except:
                quantity=""
            # print(quantity)
            try:
                immediate=result['is_immediate_delivery_item']
            except:
                immediate=""
            # print(immediate)
            data={'category':"PM",'price':price,'size':size,'quantity':quantity,'immediate':immediate}
            # print(data)
            dataList.append(data)
        if count>=PMScroll:
            print("스크롤채움")
            break
        count+=1
        time.sleep(random.randint(10,15)*0.1)
    return dataList

def GetGMTransaction(GMScroll,token,refreshToken,headers,productNo):
    cookies = {
        'i18n_redirected': 'kr',
        'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
        'AF_SYNC': '1701661188073',
        'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
        '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
        '_fbp': 'fb.2.1701936854500.2073430007',
        '_gid': 'GA1.3.1820798567.1701936855',
        'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
        'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
        'NA_SAS': '1',
        'NVADID': '0zC0001-KB5zvCorg1i4',
        'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
        'refresh_token_cookie': refreshToken,
        'csrf_refresh_token': '9b39c907-2d44-4947-952b-83e8b2b2cf55',
        'login_type': 'email',
        '_token.local': token,
        '_refresh_token.local': refreshToken,
        'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
        'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
        'strategy': 'local',
        'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
        'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'airbridge_utm_timestamp': '1701959996934',
        '_ga': 'GA1.3.248227678.1701661186',
        '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701962031505%2C%22c%22%3A1701959994791%2C%22l%22%3A1701960231505%7D',
        'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MDIzMTUwOSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODglN0Q=',
        'wcs_bt': 's_59a6a417df3:1701960232',
        '_gat_gtag_UA_153398119_1': '1',
        'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701960232744%7D',
        '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
        '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
    }
    count=1
    dataList=[]
    while True:


        params = {
            'cursor': count,
            'per_page': '50',
            'sort': '',
            'request_key': '3e327bfd-49c7-4cfb-9f68-b6c795847f23',
        }



        try:
            response = requests.get('https://kream.co.kr/api/p/products/{}/bids'.format(productNo), params=params,cookies=cookies, headers=headers)
            print(response.text)
            results = json.loads(response.text)['items']
        except:
            print("더없다")
            break

        for result in results:
            try:
                price=result['price']
            except:
                price=""
            # print(price)
            try:
                size=extract_characters(result['option'])
            except:
                size=""
            # print(size)
            try:
                quantity=result['quantity']
            except:
                quantity=""
            # print(quantity)
            # try:
            #     immediate=result['is_immediate_delivery_item']
            # except:
            #     immediate=""
            # print(immediate)
            data={'category':"GM",'price':price,'size':size,'quantity':quantity}
            # print(data)
            dataList.append(data)
        if count >= GMScroll:
            print("스크롤채움")
            break
        count+=1
        time.sleep(random.randint(10, 15) * 0.1)
    return dataList

def GetBasicData(token,refreshToken,headers,productNo):
    # # cookies = {
    # #     'i18n_redirected': 'kr',
    # #     'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
    # #     'AF_SYNC': '1701661188073',
    # #     'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
    # #     '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
    # #     '_fbp': 'fb.2.1701936854500.2073430007',
    # #     '_gid': 'GA1.3.1820798567.1701936855',
    # #     'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
    # #     'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
    # #     'NA_SAS': '1',
    # #     'NVADID': '0zC0001-KB5zvCorg1i4',
    # #     'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
    # #     'refresh_token_cookie': refreshToken,
    # #     'csrf_refresh_token': '9b39c907-2d44-4947-952b-83e8b2b2cf55',
    # #     'login_type': 'email',
    # #     '_token.local': token,
    # #     '_refresh_token.local': refreshToken,
    # #     'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
    # #     'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
    # #     'strategy': 'local',
    # #     'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
    # #     'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
    # #     'airbridge_utm_timestamp': '1701959996934',
    # #     '_ga': 'GA1.3.248227678.1701661186',
    # #     '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
    # #     'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701962031505%2C%22c%22%3A1701959994791%2C%22l%22%3A1701960231505%7D',
    # #     'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MDIzMTUwOSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODglN0Q=',
    # #     'wcs_bt': 's_59a6a417df3:1701960232',
    # #     '_gat_gtag_UA_153398119_1': '1',
    # #     'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701960232744%7D',
    # #     '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
    # #     '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701960233.59.0.0',
    # # }
    #
    # headers = {
    #     'authority': 'www.kream.co.kr',
    #     'accept': 'application/json, text/plain, */*',
    #     'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
    #     'authorization': 'Bearer {}'.format(token),
    #     # 'cookie': 'i18n_redirected=kr; did=dcee2e4a-b76e-4492-bcd2-894f2431eba9; _fbp=fb.2.1697370550036.1396063279; airbridge_device_alias=%7B%22amplitude_device_id%22%3A%2253902801-c785-4f54-96a5-ffb846bd52d1%22%7D; _gid=GA1.3.1356249296.1697370551; afUserId=9b7fcd22-3ca0-481c-9b49-c9ae4043c15e-p; AF_SYNC=1697370552206; AMP_MKTG_487619ef1d=JTdCJTIycmVmZXJyZXIlMjIlM0ElMjJodHRwcyUzQSUyRiUyRm5pZC5uYXZlci5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIybmlkLm5hdmVyLmNvbSUyMiU3RA==; _token.social_naver=false; _refresh_token.social_naver=false; refresh_token_cookie=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTY5NzM3MDU1NCwianRpIjoiNDczMWE1MjQtMDY3YS00NmQxLTkwZTktN2ViN2VhZWU5OTU3IiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NTc4NjMxMiwibmJmIjoxNjk3MzcwNTU0LCJjc3JmIjoiYjQ5ZGZkY2EtNzZhMC00Y2Q2LWI2MWMtM2FmODIwYzE5NWE0IiwiZXhwIjoxNjk3NDU2OTU0LCJ1YyI6eyJzYWZlIjp0cnVlfX0.5VxnAUR-ci7wz0RcW9V2iuJbeyCiLo9hFLZyNWVwZL8; csrf_refresh_token=b49dfdca-76a0-4cd6-b61c-3af820c195a4; login_type=social; _token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6dHJ1ZSwiaWF0IjoxNjk3MzcwNTU0LCJqdGkiOiJiN2UwNTI5ZC01ZTBkLTQzMWQtOGRiZS1kNDZmZTZiNmQ1YWQiLCJ0eXBlIjoiYWNjZXNzIiwiaWRlbnRpdHkiOjU3ODYzMTIsIm5iZiI6MTY5NzM3MDU1NCwiY3NyZiI6IjA3MzRmZjZlLWRiOTctNDgxYy05MGEyLThlZWRkYTYyOGZmZiIsImV4cCI6MTY5NzM3Nzc1NCwidWMiOnsic2FmZSI6dHJ1ZX19.mZyA8gYSjymwdjWRGUpP6S1XZ9urTLXJygJWSmOKfPY; _refresh_token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTY5NzM3MDU1NCwianRpIjoiNDczMWE1MjQtMDY3YS00NmQxLTkwZTktN2ViN2VhZWU5OTU3IiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NTc4NjMxMiwibmJmIjoxNjk3MzcwNTU0LCJjc3JmIjoiYjQ5ZGZkY2EtNzZhMC00Y2Q2LWI2MWMtM2FmODIwYzE5NWE0IiwiZXhwIjoxNjk3NDU2OTU0LCJ1YyI6eyJzYWZlIjp0cnVlfX0.5VxnAUR-ci7wz0RcW9V2iuJbeyCiLo9hFLZyNWVwZL8; strategy=local; ab.storage.userId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%2200bbc1a3-69e7-461b-8d05-34353f86ace0%22%2C%22c%22%3A1697370557824%2C%22l%22%3A1697370557825%7D; ab.storage.deviceId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%22cdaa485b-de21-e5c5-8a8d-9732b2dc23cc%22%2C%22c%22%3A1697370557826%2C%22l%22%3A1697370557826%7D; ab.storage.sessionId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%22bbecc6e7-153b-daa2-112e-1f7ff130e9ea%22%2C%22e%22%3A1697372357829%2C%22c%22%3A1697370557825%2C%22l%22%3A1697370557829%7D; wcs_bt=s_59a6a417df3:1697370557; airbridge_session=%7B%22id%22%3A%22cc5b9a55-7e70-4201-a5b9-b35ce164d253%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1697370550466%2C%22end%22%3A1697370558041%7D; _ga=GA1.3.379536745.1697370550; _ga_SRFKTMTR0R=GS1.1.1697370550.1.1.1697371318.60.0.0; _ga_5LYDPM15LW=GS1.1.1697370550.1.1.1697371318.60.0.0; AMP_487619ef1d=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjI1MzkwMjgwMS1jNzg1LTRmNTQtOTZhNS1mZmI4NDZiZDUyZDElMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjIwMGJiYzFhMy02OWU3LTQ2MWItOGQwNS0zNDM1M2Y4NmFjZTAlMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNjk3MzcwNTQ5NjMyJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTY5NzM3MDU1NTk3NSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMyU3RA==',
    #     'referer': 'https://www.kream.co.kr/products/15248',
    #     'sec-ch-ua': '"Chromium";v="118", "Google Chrome";v="118", "Not=A?Brand";v="99"',
    #     'sec-ch-ua-mobile': '?0',
    #     'sec-ch-ua-platform': '"Windows"',
    #     'sec-fetch-dest': 'empty',
    #     'sec-fetch-mode': 'cors',
    #     'sec-fetch-site': 'same-origin',
    #     'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36',
    #     'x-kream-api-version': '25',
    #     # 'x-kream-client-datetime': '20231015210158+0900',
    #     # 'x-kream-device-id': 'web;dcee2e4a-b76e-4492-bcd2-894f2431eba9',
    #     'x-kream-web-build-version': '4.17.0',
    # }
    #
    # response = requests.get('https://www.kream.co.kr/products/{}'.format(productNo),
    #                         # cookies=cookies,
    #                         headers=headers)
    #
    # soup=BeautifulSoup(response.text,'lxml')
    #
    # infoBoxs=soup.find_all('div',attrs={'class':re.compile('detail-box+')})
    # for infoBox in infoBoxs:
    #     if infoBox.get_text().find("발매가")>=0:
    #         originPrice=infoBox.find('div',attrs={'class':'product_info'}).get_text().replace(",","").replace("원","")
    #         print("originPrice:",originPrice)
    #     elif infoBox.get_text().find("모델번호")>=0:
    #         modelCode=infoBox.find('div',attrs={'class':'product_info'}).get_text().strip()
    #         print('modelCode:',modelCode)
    # try:
    #     originPrice=int(originPrice)
    # except:
    #     originPrice=99999999
    #     print("숫자변환불가")
    # try:
    #     title = soup.find('p', attrs={'class': 'sub-title'}).get_text()
    # except:
    #     title = ""
    # print("title:",title)
    #
    # try:
    #     imageUrl=soup.find('picture',attrs={'class':'picture product_img'}).find('source')['srcset']
    # except:
    #     imageUrl=""
    # print("imageUrl:",imageUrl)
    #
    # try:
    #     url='https://www.kream.co.kr/products/{}'.format(productNo)
    # except:
    #     url=""
    # print('url:',url)
    # data={'modelCode':modelCode,'originPrice':originPrice,'title':title,'imageUrl':imageUrl,'url':url}
    # return data
    #=====================과거코드
    cookies = {
        'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
        'AF_SYNC': '1701661188073',
        'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
        'i18n_redirected': 'kr',
        '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
        '_fbp': 'fb.2.1701936854500.2073430007',
        '_gid': 'GA1.3.1820798567.1701936855',
        'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
        'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
        'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
        'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'airbridge_utm_timestamp': '1701959996934',
        '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'strategy': 'local',
        'did': 'd0809d58-4cae-4ea1-a8fd-d71d66119564',
        'AMP_MKTG_487619ef1d': 'JTdCJTdE',
        '_token.local': 'false',
        'login_type': 'email',
        'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%228b983ca8-c9e0-c5b6-762a-b068ad93c824%22%2C%22e%22%3A1702131758394%2C%22c%22%3A1702129950535%2C%22l%22%3A1702129958394%7D',
        'wcs_bt': 's_59a6a417df3:1702129968',
        '_ga': 'GA1.3.248227678.1701661186',
        'airbridge_session': '%7B%22id%22%3A%22bde47340-9f67-48ca-a5d4-535b70c752be%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1702129950182%2C%22end%22%3A1702129968643%7D',
        '_ga_SRFKTMTR0R': 'GS1.1.1702129949.13.1.1702130081.60.0.0',
        '_ga_5LYDPM15LW': 'GS1.1.1702129949.13.1.1702130081.60.0.0',
        'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAyMTI5OTUwNTMxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMjEyOTk1ODQwNSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMTExJTdE',
    }

    headers = {
        'authority': 'www.kream.co.kr',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        # 'cookie': 'afUserId=92d9569b-f0e5-4a77-aed1-f0b3ac529581-p; AF_SYNC=1701661188073; airbridge_device_alias=%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D; i18n_redirected=kr; _fwb=11Uusi3FTaGqJopqZc3EJx.1701833305082; _fbp=fb.2.1701936854500.2073430007; _gid=GA1.3.1820798567.1701936855; ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D; ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D; airbridge_utm=%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D; airbridge_utm_url=https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; airbridge_utm_timestamp=1701959996934; _gac_UA-153398119-1=1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; strategy=local; did=d0809d58-4cae-4ea1-a8fd-d71d66119564; AMP_MKTG_487619ef1d=JTdCJTdE; _token.local=false; login_type=email; ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%228b983ca8-c9e0-c5b6-762a-b068ad93c824%22%2C%22e%22%3A1702131758394%2C%22c%22%3A1702129950535%2C%22l%22%3A1702129958394%7D; wcs_bt=s_59a6a417df3:1702129968; _ga=GA1.3.248227678.1701661186; airbridge_session=%7B%22id%22%3A%22bde47340-9f67-48ca-a5d4-535b70c752be%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1702129950182%2C%22end%22%3A1702129968643%7D; _ga_SRFKTMTR0R=GS1.1.1702129949.13.1.1702130081.60.0.0; _ga_5LYDPM15LW=GS1.1.1702129949.13.1.1702130081.60.0.0; AMP_487619ef1d=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAyMTI5OTUwNTMxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMjEyOTk1ODQwNSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMTExJTdE',
        'referer': 'https://www.kream.co.kr/products/84285',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'x-kream-api-version': '25',
        'x-kream-client-datetime': '20231209225442+0900',
        'x-kream-device-id': 'web;d0809d58-4cae-4ea1-a8fd-d71d66119564',
        'x-kream-web-build-version': '4.17.0',
    }

    params = {
        'sort': 'featured',
        'per_page': '12',
        'with_recent_comments': 'true',
        'request_key': '906fc9e3-494d-4889-81cf-ee071a5654cc',
    }

    response = requests.get('https://www.kream.co.kr/api/s/feed/products/{}/'.format(str(productNo)), params=params, cookies=cookies,
                            headers=headers)
    results = json.loads(response.text)
    findResults = find_values_by_key2(results, 'release')
    # pprint.pprint(findResults)

    # id 값이 84285인 요소 찾기
    found_item = next((item for item in findResults if item.get("id") == int(productNo)), None)

    # 결과 출력
    if found_item:
        print(found_item)
    else:
        print("해당 id를 가진 요소가 없습니다.")

    try:
        title = found_item['translated_name']
    except:
        title = ""
    print("title:", title)

    url = 'https://www.kream.co.kr/products/{}'.format(productNo)

    try:
        originPrice = found_item['original_price']
    except:
        originPrice = ""
    print("originPrice:", originPrice)

    try:
        imageUrl = found_item['image_urls'][0]
    except:
        imageUrl = ""
    print("imageUrl:", imageUrl)

    try:
        modelCode = found_item['style_code']
    except:
        modelCode = ""
    print("modelCode:", modelCode)
    data = {'modelCode': modelCode, 'originPrice': originPrice, 'title': title, 'imageUrl': imageUrl, 'url': url}
    return data

def GetSearch(fname,inputValues):
    wb=openpyxl.load_workbook(fname)
    sheetNames=wb.sheetnames
    print(sheetNames)

    xb=openpyxl.Workbook()
    xs=xb.active
    # columnName=['사진URL','URL','품명','모델번호','체결사이즈','전체거래가평균(O)', '전체 거래가 평균(X)', '전체 거래가 평균(O+X)', '3개 까지의 거래가 평균(X)/전체 거래가 평균(X)',
    #               '전체 거래가 평균(O)/전체 거래가 평균(X)(=B/C)', '체결거래수(O+X)', '구매입찰 개수/판매입찰 개수', '거래가 평균(X)/매도 호가', '판매 호가(첫 행값)/발매가',
    #               '판매입찰 개수', '거래가 평균(X)/매수 호가', '구매 호가(첫 행값) / 발매가', '구매입찰개수']
    columnName=['사진URL',
     'URL',
     '체결사이즈',
     '전체 거래가 평균(X)',
     '품명',
     '모델번호',
     '전체거래가평균(O)',
     '전체 거래가 평균(O+X)',
     '3개 까지의 거래가 평균(X)/전체 거래가 평균(X)',
     '전체 거래가 평균(O)/전체 거래가 평균(X)(=B/C)',
     '체결거래수(O+X)',
     '구매입찰 개수/판매입찰 개수',
     '거래가 평균(X)/매도 호가',
     '판매 호가(첫 행값)/발매가',
     '판매입찰 개수',
     '거래가 평균(X)/매수 호가',
     '구매 호가(첫 행값) / 발매가',
     '구매입찰개수']
    xs.append(columnName)
    timeNow=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    print("1234")
    for sheetName in sheetNames:
        print("===================================================")
        print("시트이름은:",sheetName)
        ws=wb[sheetName]
        valueB = ws.cell(row=2, column=2).value
        valueC = ws.cell(row=2, column=3).value
        valueD = ws.cell(row=2, column=4).value
        valueE = ws.cell(row=2, column=5).value
        valueF = ws.cell(row=2, column=6).value
        valueG = ws.cell(row=2, column=7).value
        valueH = ws.cell(row=2, column=8).value
        valueI = ws.cell(row=2, column=9).value
        valueJ = ws.cell(row=2, column=10).value
        valueK = ws.cell(row=2, column=11).value
        valueL = ws.cell(row=2, column=12).value
        valueM = ws.cell(row=2, column=13).value
        valueN = ws.cell(row=2, column=14).value
        if valueB==None:
            continue
        if valueC==None:
            continue
        if valueD==None:
            continue
        if valueE==None:
            continue
        if valueF==None:
            continue
        if valueG==None:
            continue
        if valueH==None:
            continue
        if valueI==None:
            continue
        if valueJ==None:
            continue
        if valueK==None:
            continue
        if valueL==None:
            continue
        if valueM==None:
            continue
        if valueN==None:
            continue
        myValue={'valueD':valueD,'valueE':valueE,'valueF':valueF,'valueG':valueG,'valueH':valueH,'valueI':valueI,'valueJ':valueJ,'valueK':valueK,'valueL':valueL,'valueM':valueM,'valueN':valueN}
        print('myValue:',myValue)
        checkResult=[]
        for inputValue in inputValues:
            print(inputValue)
            if inputValue['switch']=="UP":
                print(myValue[inputValue['name']],type(myValue[inputValue['name']]),inputValue['value'],type(inputValue['value']))
                if float(myValue[inputValue['name']])>=float(inputValue['value']):
                    checkResult.append(True)
                else:
                    checkResult.append(False)
            if inputValue['switch']=="DOWN":
                if float(myValue[inputValue['name']])<=float(inputValue['value']):
                    checkResult.append(True)
                else:
                    checkResult.append(False)
        print('checkResult:',checkResult)

        allTrue = all(checkResult)

        # 결과 출력
        if allTrue:
            print('찾았다')
            url=ws.cell(row=2,column=1).value
            imageUrl = ws.cell(row=4, column=1).value
            modelName = ws.cell(row=4, column=3).value
            modelCode=ws.cell(row=4, column=4).value
            size = ws.cell(row=4, column=5).value
            # data=[imageUrl,url,modelName,modelCode,size,valueB,valueC,valueD,valueE,valueF,valueG,valueH,valueI,valueJ,valueK,valueL,valueM,valueN]
            data=[imageUrl,url,size,valueC,modelName,modelCode,valueB,valueD,valueE,valueF,valueG,valueH,valueI,valueJ,valueK,valueL,valueM,valueN]
            regiDate=datetime.datetime.now().strftime("%Y%m%d %H:%M:%S")
            regiTimestamp=datetime.datetime.now().timestamp()
            try:
                headers = {
                    'accept': 'application/json',
                    'Content-Type': 'application/json',
                }

                json_data = {
                    'parameter': {
                        'id': 0,
                        'imageUrl': imageUrl,
                        'url': url,
                        'size': size,
                        'avgX': valueC,
                        'productName': modelName,
                        'modelNumber': modelCode,
                        'avgO': valueB,
                        'avgOX': valueD,
                        'avg3PeravgX': valueE,
                        'avgOPerAvgX': valueF,
                        'tradeCountOX': valueG,
                        'buyCountPersellCount': valueH,
                        'avgXPerSellPrice': valueI,
                        'sellPricePerOriginPrice': valueJ,
                        'sellCount': valueK,
                        'avgXPerBuyPrice': valueL,
                        'buyPricePerOriginPrice': valueM,
                        'buyCount': valueN,
                        'regiDate':regiDate,
                        'regiTimestamp':regiTimestamp
                    },
                    # valueD, valueE, valueF, valueG, valueH, valueI, valueJ, valueK, valueL, valueM, valueN
                }
                response = requests.post('https://omgsapzmjvhh7o7ng6ewb65iyy0fxcty.lambda-url.ap-northeast-2.on.aws/product/create', headers=headers, json=json_data)
                print("전송완료")
            except:
                print("전송불가")
            xs.append(data)

            # 첫 번째 행 고정
            xs.freeze_panes = 'A2'

            column_dimension="A:R"
            # # 전체 열에 필터 적용
            xs.auto_filter.ref = column_dimension
            searchfilename='search_result_{}.xlsx'.format(timeNow)
            xb.save(searchfilename)
        else:
            print("해당없음")
    return searchfilename

def GetGoogleSpreadSheet():
    scope = 'https://spreadsheets.google.com/feeds'
    json = 'credential_client.json'
    credentials = ServiceAccountCredentials.from_json_keyfile_name(json, scope)
    gc = gspread.authorize(credentials)
    sheet_url = 'https://docs.google.com/spreadsheets/d/1xd0R6OePv6mc78GPgkz8nKwXXsH44AHxJ-plFTpZC6I/edit?hl=ko&pli=1#gid=0'
    doc = gc.open_by_url(sheet_url)
    worksheet = doc.worksheet('시트1')
    #=================특정행의 정보 가져오기
    # cell_data = worksheet.acell('A1').value
    #=================전체정보가져오기
    # all_data=worksheet.get_all_records()
    # print("all_data:",all_data,"/ all_data_TYPE:",type(all_data))

    # 열 이름 가져오기 (첫 번째 행의 값으로 가정)
    header_row = worksheet.row_values(1)

    # 각 열의 데이터를 딕셔너리로 저장
    data_dict = {}
    for col_index, col_name in enumerate(header_row):
        data = worksheet.col_values(col_index + 1)[1:]  # 첫 번째 행은 열 이름이므로 제외
        data_dict[col_name] = data
    print(data_dict)
    return data_dict

    #==================맨 밑행에 데이타 넣기

def GetToken():
    cookies = {
        'i18n_redirected': 'kr',
        'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
        'AF_SYNC': '1701661188073',
        'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
        '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
        '_fbp': 'fb.2.1701936854500.2073430007',
        '_gid': 'GA1.3.1820798567.1701936855',
        'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701936863775%7D',
        'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
        '_token.local': 'false',
        'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
        'NA_SAS': '1',
        'NVADID': '0zC0001-KB5zvCorg1i4',
        '_refresh_token.local': 'false',
        'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
        '_gac_UA-153398119-1': '1.1701959965.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        '_gat_gtag_UA_153398119_1': '1',
        'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
        'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
        'airbridge_utm_timestamp': '1701959965224',
        'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22eb15a355-54eb-f80a-9e2b-b64386e56a98%22%2C%22e%22%3A1701961777465%2C%22c%22%3A1701959977466%2C%22l%22%3A1701959977466%7D',
        'wcs_bt': 's_59a6a417df3:1701959977',
        'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjJjZWRmNjk1ZC1jMzJhLTQxNjctOWJmNi0wNjFlOGI0NTdmNTclMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk1OTk3NzQ3NSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODUlN0Q=',
        'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701959977548%7D',
        '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701959977.47.0.0',
        '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701959977.47.0.0',
        '_ga': 'GA1.3.248227678.1701661186',
    }

    headers = {
        'authority': 'kream.co.kr',
        'accept': 'application/json, text/plain, */*',
        'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
        'content-type': 'application/json',
        # 'cookie': 'i18n_redirected=kr; afUserId=92d9569b-f0e5-4a77-aed1-f0b3ac529581-p; AF_SYNC=1701661188073; airbridge_device_alias=%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D; _fwb=11Uusi3FTaGqJopqZc3EJx.1701833305082; _fbp=fb.2.1701936854500.2073430007; _gid=GA1.3.1820798567.1701936855; ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701936863775%7D; did=44c3f570-2970-47b3-9fc2-a23f225698eb; _token.local=false; NA_SA=Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==; NA_SAS=1; NVADID=0zC0001-KB5zvCorg1i4; _refresh_token.local=false; AMP_MKTG_487619ef1d=JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=; _gac_UA-153398119-1=1.1701959965.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; _gat_gtag_UA_153398119_1=1; airbridge_utm=%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D; airbridge_utm_url=https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; airbridge_utm_timestamp=1701959965224; ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%22eb15a355-54eb-f80a-9e2b-b64386e56a98%22%2C%22e%22%3A1701961777465%2C%22c%22%3A1701959977466%2C%22l%22%3A1701959977466%7D; wcs_bt=s_59a6a417df3:1701959977; AMP_487619ef1d=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjJjZWRmNjk1ZC1jMzJhLTQxNjctOWJmNi0wNjFlOGI0NTdmNTclMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk1OTk3NzQ3NSUyQyUyMmxhc3RFdmVudElkJTIyJTNBODUlN0Q=; airbridge_session=%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701959977548%7D; _ga_SRFKTMTR0R=GS1.1.1701959964.8.1.1701959977.47.0.0; _ga_5LYDPM15LW=GS1.1.1701959964.8.1.1701959977.47.0.0; _ga=GA1.3.248227678.1701661186',
        'origin': 'https://kream.co.kr',
        'referer': 'https://kream.co.kr/login',
        'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'sec-fetch-dest': 'empty',
        'sec-fetch-mode': 'cors',
        'sec-fetch-site': 'same-origin',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'x-kream-api-version': '25',
        'x-kream-client-datetime': '20231207233954+0900',
        'x-kream-device-id': 'web;44c3f570-2970-47b3-9fc2-a23f225698eb',
        'x-kream-web-build-version': '4.17.0',
    }

    # params = {
    #     'request_key': '039854fd-67ea-443d-a2b9-065d71bb2ba0',
    # }
    #대표님 테스트계정
    # json_data = {
    #     'email': 'zmflal001@gmail.com',
    #     'password': 'zmflal001*',
    # }
    #대표님 서버 계정
    json_data = {
        'email': 'skytogether7',
        'password': 'Qotpgksdl1!',
    }

    response = requests.post('https://kream.co.kr/api/auth/login', cookies=cookies, headers=headers,
                             json=json_data)
    print(response.text)
    results=json.loads(response.text)
    with open('tokenData.json', 'w',encoding='utf-8-sig') as f:
        json.dump(results, f, indent=2,ensure_ascii=False)
    return results

def GetIDs(token,refreshToken,catId,noLimit):
    pageNo=1
    productNoList=[]
    endFlag=False
    while True:
        cookies = {
            'i18n_redirected': 'kr',
            'afUserId': '92d9569b-f0e5-4a77-aed1-f0b3ac529581-p',
            'AF_SYNC': '1701661188073',
            'airbridge_device_alias': '%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D',
            '_fwb': '11Uusi3FTaGqJopqZc3EJx.1701833305082',
            '_fbp': 'fb.2.1701936854500.2073430007',
            '_gid': 'GA1.3.1820798567.1701936855',
            'did': '44c3f570-2970-47b3-9fc2-a23f225698eb',
            'NA_SA': 'Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==',
            'NA_SAS': '1',
            'NVADID': '0zC0001-KB5zvCorg1i4',
            'AMP_MKTG_487619ef1d': 'JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=',
            'refresh_token_cookie': refreshToken,
            'csrf_refresh_token': '9b39c907-2d44-4947-952b-83e8b2b2cf55',
            'login_type': 'email',
            '_token.local': token,
            '_refresh_token.local': refreshToken,
            'ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D',
            'ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D',
            'strategy': 'local',
            'airbridge_utm': '%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D',
            'airbridge_utm_url': 'https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
            'airbridge_utm_timestamp': '1701959996934',
            '_ga': 'GA1.3.248227678.1701661186',
            '_gac_UA-153398119-1': '1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE',
            '_gat_gtag_UA_153398119_1': '1',
            'ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37': '%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701964093692%2C%22c%22%3A1701959994791%2C%22l%22%3A1701962293692%7D',
            'AMP_487619ef1d': 'JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MjI5MzcwMiUyQyUyMmxhc3RFdmVudElkJTIyJTNBOTMlN0Q=',
            'wcs_bt': 's_59a6a417df3:1701962295',
            'airbridge_session': '%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701962295279%7D',
            '_ga_SRFKTMTR0R': 'GS1.1.1701959964.8.1.1701962296.12.0.0',
            '_ga_5LYDPM15LW': 'GS1.1.1701959964.8.1.1701962296.12.0.0',
        }

        headers = {
            'authority': 'kream.co.kr',
            'accept': 'application/json, text/plain, */*',
            'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
            # 'authorization': 'Bearer {}'.format(token),
            # 'cookie': 'i18n_redirected=kr; afUserId=92d9569b-f0e5-4a77-aed1-f0b3ac529581-p; AF_SYNC=1701661188073; airbridge_device_alias=%7B%22amplitude_device_id%22%3A%22ffdb56da-7661-42f0-8d87-4d1d1b59d46c%22%7D; _fwb=11Uusi3FTaGqJopqZc3EJx.1701833305082; _fbp=fb.2.1701936854500.2073430007; _gid=GA1.3.1820798567.1701936855; did=44c3f570-2970-47b3-9fc2-a23f225698eb; NA_SA=Y2k9MHpDMDAwMS1LQjV6dkNvcmcxaTR8dD0xNzAxOTU0MTI1fHU9aHR0cHMlM0ElMkYlMkZrcmVhbS5jby5rciUyRiUzRnV0bV9zb3VyY2UlM0RuYXZlciUyNnV0bV9tZWRpdW0lM0RjcG0lMjZ1dG1fY2FtcGFpZ24lM0RCUyUyNnV0bV9ncm91cCUzRFBDXzIzMTIwNCUyNnV0bV9jb250ZW50JTNEaG9tZWxpbmslMjZ1dG1fdGVybSUzREtSRUFNJTI2bl9tZWRpYSUzRDI3NzU4JTI2bl9xdWVyeSUzREtSRUFNJTI2bl9yYW5rJTNEMSUyNm5fYWRfZ3JvdXAlM0RncnAtYTAwMS0wNC0wMDAwMDAwMzY5Nzk0ODElMjZuX2FkJTNEbmFkLWEwMDEtMDQtMDAwMDAwMjc1OTQ3NDg3JTI2bl9rZXl3b3JkX2lkJTNEbmt3LWEwMDEtMDQtMDAwMDA1NjE5NTA1NzkzJTI2bl9rZXl3b3JkJTNES1JFQU0lMjZuX2NhbXBhaWduX3R5cGUlM0Q0JTI2bl9jb250cmFjdCUzRHRjdC1hMDAxLTA0LTAwMDAwMDAwMDgwNjMxOCUyNm5fYWRfZ3JvdXBfdHlwZSUzRDUlMjZOYVBtJTNEY3QlMjUzRGxwdjdqaDh3JTI1N0NjaSUyNTNEMHpDMDAwMS1LQjV6dkNvcmcxaTQlMjU3Q3RyJTI1M0Ricm5kJTI1N0NoayUyNTNEODUyMGMxMjc1MTlkZDQyZTA3OWMyNzMxOGZlOGFlM2RkMzNiMzE3ZnxyPWh0dHBzJTNBJTJGJTJGc2VhcmNoLm5hdmVyLmNvbSUyRnNlYXJjaC5uYXZlciUzRndoZXJlJTNEbmV4ZWFyY2glMjZzbSUzRHRvcF9odHklMjZmYm0lM0QwJTI2aWUlM0R1dGY4JTI2cXVlcnklM0RrcmVhbQ==; NA_SAS=1; NVADID=0zC0001-KB5zvCorg1i4; AMP_MKTG_487619ef1d=JTdCJTIydXRtX2NhbXBhaWduJTIyJTNBJTIyTkVXXyVFQyU5RSU5MCVFQyU4MiVBQyVFQiVBQSU4NV8lRUMlODglOTglRUIlOEYlOTlfUEMlMjIlMkMlMjJ1dG1fY29udGVudCUyMiUzQSUyMkEuJTIwJUVDJTlFJTkwJUVDJTgyJUFDJUVCJUFBJTg1XyVFQyU4OCU5OCVFQiU4RiU5OSUyMiUyQyUyMnV0bV9tZWRpdW0lMjIlM0ElMjJjcGMlMjIlMkMlMjJ1dG1fc291cmNlJTIyJTNBJTIyZ29vZ2xlJTIyJTJDJTIydXRtX3Rlcm0lMjIlM0ElMjIlRUQlODElQUMlRUIlQTYlQkMlMjIlMkMlMjJyZWZlcnJlciUyMiUzQSUyMmh0dHBzJTNBJTJGJTJGd3d3Lmdvb2dsZS5jb20lMkYlMjIlMkMlMjJyZWZlcnJpbmdfZG9tYWluJTIyJTNBJTIyd3d3Lmdvb2dsZS5jb20lMjIlMkMlMjJnY2xpZCUyMiUzQSUyMkNqd0tDQWlBOThXckJoQVlFaXdBMld2aE9pRTF6cnlVdjZaUkl2Y2QwQmhlalowRTRyeUtKM3dxTzVqcUt0R2ZER1VHbVQ2VzFDQURJaG9DUGhrUUF2RF9Cd0UlMjIlN0Q=; refresh_token_cookie=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTcwMTk1OTk5MywianRpIjoiZjcyMTcxZWItNGE5OC00NjU1LTg4NDUtMzZmNWFiN2Q3MjFkIiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NDYyMzM1NSwibmJmIjoxNzAxOTU5OTkzLCJjc3JmIjoiOWIzOWM5MDctMmQ0NC00OTQ3LTk1MmItODNlOGIyYjJjZjU1IiwiZXhwIjoxNzAyMDQ2MzkzLCJ1YyI6eyJzYWZlIjp0cnVlfX0.9l6Q3f2CvxrC3LW6a8VG6_XrCIQBxbjon2OX6g55Ax0; csrf_refresh_token=9b39c907-2d44-4947-952b-83e8b2b2cf55; login_type=email; _token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6dHJ1ZSwiaWF0IjoxNzAxOTU5OTkzLCJqdGkiOiJjMDc5ZDg0ZC00MTYyLTRkNjItYTNlYi0xZDU3M2IyOWQyNmQiLCJ0eXBlIjoiYWNjZXNzIiwiaWRlbnRpdHkiOjQ2MjMzNTUsIm5iZiI6MTcwMTk1OTk5MywiY3NyZiI6ImM2NzUwZTczLTUwYWMtNGJkMS1hODVkLWY5MWE2YTZmZDQ0NyIsImV4cCI6MTcwMTk2NzE5MywidWMiOnsic2FmZSI6dHJ1ZX19.O4MSsTZkqaQw31svHtE4WgaTIv0RHtSQ_Mlx-4-QvnE; _refresh_token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTcwMTk1OTk5MywianRpIjoiZjcyMTcxZWItNGE5OC00NjU1LTg4NDUtMzZmNWFiN2Q3MjFkIiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NDYyMzM1NSwibmJmIjoxNzAxOTU5OTkzLCJjc3JmIjoiOWIzOWM5MDctMmQ0NC00OTQ3LTk1MmItODNlOGIyYjJjZjU1IiwiZXhwIjoxNzAyMDQ2MzkzLCJ1YyI6eyJzYWZlIjp0cnVlfX0.9l6Q3f2CvxrC3LW6a8VG6_XrCIQBxbjon2OX6g55Ax0; ab.storage.deviceId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%22e10b0036-816e-3e47-fb0a-e136dbfadb43%22%2C%22c%22%3A1701936863775%2C%22l%22%3A1701959994791%7D; ab.storage.userId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%226366f46a-b525-46b6-9e07-af6dcba56c49%22%2C%22c%22%3A1701959994789%2C%22l%22%3A1701959994791%7D; strategy=local; airbridge_utm=%7B%22channel%22%3A%22google%22%2C%22parameter%22%3A%7B%22medium%22%3A%22cpc%22%2C%22campaign%22%3A%22NEW_%uC790%uC0AC%uBA85_%uC218%uB3D9_PC%22%2C%22term%22%3A%22%uD06C%uB9BC%22%2C%22content%22%3A%22A.%20%uC790%uC0AC%uBA85_%uC218%uB3D9%22%7D%7D; airbridge_utm_url=https%3A//kream.co.kr/%3Futm_source%3Dgoogle%26utm_medium%3Dcpc%26utm_campaign%3DNEW_%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599_PC%26utm_term%3D%25ED%2581%25AC%25EB%25A6%25BC%26utm_content%3DA.%2520%25EC%259E%2590%25EC%2582%25AC%25EB%25AA%2585_%25EC%2588%2598%25EB%258F%2599%26gclid%3DCjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; airbridge_utm_timestamp=1701959996934; _ga=GA1.3.248227678.1701661186; _gac_UA-153398119-1=1.1701959997.CjwKCAiA98WrBhAYEiwA2WvhOiE1zryUv6ZRIvcd0BhejZ0E4ryKJ3wqO5jqKtGfDGUGmT6W1CADIhoCPhkQAvD_BwE; _gat_gtag_UA_153398119_1=1; ab.storage.sessionId.a45e842b-5d46-46bf-8f41-2f75d6fd4b37=%7B%22g%22%3A%225589cdf9-8744-d52e-0947-d98aae11f215%22%2C%22e%22%3A1701964093692%2C%22c%22%3A1701959994791%2C%22l%22%3A1701962293692%7D; AMP_487619ef1d=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjJmZmRiNTZkYS03NjYxLTQyZjAtOGQ4Ny00ZDFkMWI1OWQ0NmMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjI2MzY2ZjQ2YS1iNTI1LTQ2YjYtOWUwNy1hZjZkY2JhNTZjNDklMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNzAxOTU5OTYzMDgxJTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTcwMTk2MjI5MzcwMiUyQyUyMmxhc3RFdmVudElkJTIyJTNBOTMlN0Q=; wcs_bt=s_59a6a417df3:1701962295; airbridge_session=%7B%22id%22%3A%226ec19915-5bc6-464b-afe0-7307b9958696%22%2C%22timeout%22%3A1800000%2C%22start%22%3A1701959965244%2C%22end%22%3A1701962295279%7D; _ga_SRFKTMTR0R=GS1.1.1701959964.8.1.1701962296.12.0.0; _ga_5LYDPM15LW=GS1.1.1701959964.8.1.1701962296.12.0.0',
            'referer': 'https://kream.co.kr/search?shop_category_id=34&sort=popularity_without_trading',
            'sec-ch-ua': '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'x-kream-api-version': '25',
            'x-kream-client-datetime': '20231208001818+0900',
            'x-kream-device-id': 'web;44c3f570-2970-47b3-9fc2-a23f225698eb',
            'x-kream-web-build-version': '4.17.0',
        }

        params = {
            'cursor': pageNo,
            'shop_category_id': catId,
            'sort': 'popularity_without_trading',
            'request_key': '96c769cd-3491-4ff8-9960-9be2ba011182',
        }

        response = requests.get('https://kream.co.kr/api/p/tabs/all/', params=params, cookies=cookies, headers=headers)

        results=json.loads(response.text)['items']

        # pprint.pprint(results)
        idList=[]
        if len(results)==0:
            break
        for result in results:
            # pprint.pprint(results)
            try:
                productIds=result['product']['release']['id']
                print("productIds:",productIds,"/ productIds_TYPE:",type(productIds))
                productNoList.append(productIds)
            except:
                print("없음")

            if len(productNoList)>=noLimit:
                endFlag=True
                break
            print("====================")

        with open('productNoList.json', 'w',encoding='utf-8-sig') as f:
            json.dump(productNoList, f, indent=2,ensure_ascii=False)

        if endFlag==True:
            break
        pageNo+=1
        time.sleep(random.randint(5,10)*0.1)
    return productNoList

def SendMail(category,filepath):

    smtp_server = 'smtp.naver.com'
    smtp_port = 587

    # 네이버 이메일 계정 정보
    username = 'wsgt18@naver.com'  # 클라이언트 정보 입력
    password = 'dnrglvotl0*'  # 클라이언트 정보 입력

    # receiver='wsgt17@naver.com'
    receiver='wsgt17@naver.com'
    # receiver=email

    # username = 'hellfir2@naver.com'  # 클라이언트 정보 입력
    # password = 'dlwndwo1!'  # 클라이언트 정보 입력
    # =================커스터마이징
    try:
        to_mail = receiver
    except:
        print("메일주소없음")
        return

    # =================

    # 메일 수신자 정보
    to_email = receiver

    # 참조자 정보
    cc_email = 'ljj3347@naver.com'

    # 메일 본문 및 제목 설정
    contentList=[]

    content="\n".join(contentList)


    # MIMEMultipart 객체 생성
    timeNow=datetime.datetime.now().strftime("%Y년%m월%d일 %H시%M분%S초")
    msg = MIMEMultipart('alternative')
    msg["Subject"] = "[결과]크림 상품 크롤링 ({}){}".format(str(category+1)+"번째 카테고리",timeNow)  # 메일 제목
    msg['From'] = username
    msg['To'] = to_email
    msg['Cc'] = cc_email  # 참조 이메일 주소 추가
    msg.attach(MIMEText(content, 'plain'))

    # 파일 첨부
    part = MIMEBase('application', 'octet-stream')
    with open(filepath, 'rb') as file:
        part.set_payload(file.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename={filepath}')
    msg.attach(part)

    # SMTP 서버 연결 및 로그인
    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(username, password)
    # 이메일 전송 (수신자와 참조자 모두에게 전송)
    to_and_cc_emails = [to_email] + [cc_email]
    server.sendmail(username, to_and_cc_emails, msg.as_string())
    # SMTP 서버 연결 종료
    server.quit()
    print("전송완료")

# =============================

def fetchData():
    while True:
        try:
            data=GetGoogleSpreadSheet()
            print("성공")
            break
        except:
            print("실패")
        time.sleep(60)

    print("data:",data,"/ data_TYPE:",type(data))

    noLimit=int(data['상품수(카테고리그룹당)'][0])
    catIds=data['검색카테고리']
    daysAgo=int(data['체결기간'][0])
    PMScroll=int(data['판매스크롤'][0])
    GMScroll=int(data['구매스크롤'][0])
    skipCount=int(data['체결(X)스킵 기준'][0])

    #=======================로그인
    # GetToken()
    # time.sleep(2)
    # with open('tokenData.json', "r", encoding='utf-8-sig') as f:
    #     tokenData = json.load(f)




    for catIndex,catId in enumerate(catIds):
        GetToken()
        with open ('tokenData.json', "r",encoding='utf-8-sig') as f:
            tokenData = json.load(f)
        token = tokenData['access_token']
        refreshToken = tokenData['refresh_token']
        print("토큰가져오기완료1")

        GetIDs(token, refreshToken, catId, noLimit)
        with open ('productNoList.json', "r",encoding='utf-8-sig') as f:
            productNoList = json.load(f)

        #===============상세졍보가져오기

        wb = openpyxl.Workbook()
        ws = wb.active
        columnName = ['상품URL', '전체거래가평균(O)', '전체 거래가 평균(X)', '전체 거래가 평균(O+X)', '3개 까지의 거래가 평균(X)/전체 거래가 평균(X)',
                      '전체 거래가 평균(O)/전체 거래가 평균(X)(=B/C)', '체결거래수(O+X)', '구매입찰 개수/판매입찰 개수', '거래가 평균(X)/매도 호가',
                      '판매 호가(첫 행값)/발매가',
                      '판매입찰 개수', '거래가 평균(X)/매수 호가', '구매 호가(첫 행값) / 발매가', '구매입찰개수']
        ws.append(columnName)
        makeFlag = True
        timeNow = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        for index,productNo in enumerate(productNoList):
            if index%500==0 and index>=10:
                GetToken()
                with open('tokenData.json', "r", encoding='utf-8-sig') as f:
                    tokenData = json.load(f)
                token = tokenData['access_token']
                refreshToken = tokenData['refresh_token']
            while True:
                try:
                    # =================ID가져오기
                    with open('tokenData.json', "r", encoding='utf-8-sig') as f:
                        tokenData = json.load(f)
                    token = tokenData['access_token']
                    refreshToken = tokenData['refresh_token']
                    print("토큰 가져오기 완료2")
                    break
                except:
                    print("잠시에러")
                    time.sleep(1)


            print("productNo:", productNo)
            text="{}번째 상품 확인중...".format(index+1)
            print(text)
            headers = {
                'authority': 'www.kream.co.kr',
                'accept': 'application/json, text/plain, */*',
                'accept-language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                'authorization': 'Bearer {}'.format(token),
                # 'cookie': 'afUserId=cd0a42c4-1500-45e1-a551-746510b06fbf-p; _fbp=fb.2.1691128991158.731815584; i18n_redirected=kr; _gid=GA1.3.1693048652.1694520729; AF_SYNC=1694520730521; did=75448f17-2f9f-4c5d-bb93-45a642909201; AMP_MKTG_487619ef1d=JTdCJTdE; _gat_gtag_UA_153398119_1=1; _token.social_naver=false; _refresh_token.social_naver=false; refresh_token_cookie=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTY5NDcwNDE1OCwianRpIjoiMWFkNGQ5NzUtZTJlYi00NjExLWFjZDYtYTU2OTYwOTVmZmEwIiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NTc4NjMxMiwibmJmIjoxNjk0NzA0MTU4LCJjc3JmIjoiNWM2YTYzODMtOTM3My00NjlkLWIwN2YtNzZmOWIxMjQwMWJiIiwiZXhwIjoxNjk0NzkwNTU4LCJ1YyI6eyJzYWZlIjp0cnVlfX0.vOXtvTNZX4-H_qcHPgxMOXA3lEma3XPtK5Q36PiB1jg; csrf_refresh_token=5c6a6383-9373-469d-b07f-76f9b12401bb; login_type=social; _token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6dHJ1ZSwiaWF0IjoxNjk0NzA0MTU4LCJqdGkiOiJiZDI5YzYzOS02NTFlLTQ3ODktYjQ4YS1lMzUwNmI1YjhiNmQiLCJ0eXBlIjoiYWNjZXNzIiwiaWRlbnRpdHkiOjU3ODYzMTIsIm5iZiI6MTY5NDcwNDE1OCwiY3NyZiI6IjBmNjg4NzZjLWI3OGUtNGMwZi1hNTQ3LTUzNmQ2OGNlODk5MCIsImV4cCI6MTY5NDcxMTM1OCwidWMiOnsic2FmZSI6dHJ1ZX19.YHdNCKgtZgsaOFj89S4ZZBhquXVGyzJx1pOhNCeBNTc; _refresh_token.local=eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJmcmVzaCI6ZmFsc2UsImlhdCI6MTY5NDcwNDE1OCwianRpIjoiMWFkNGQ5NzUtZTJlYi00NjExLWFjZDYtYTU2OTYwOTVmZmEwIiwidHlwZSI6InJlZnJlc2giLCJpZGVudGl0eSI6NTc4NjMxMiwibmJmIjoxNjk0NzA0MTU4LCJjc3JmIjoiNWM2YTYzODMtOTM3My00NjlkLWIwN2YtNzZmOWIxMjQwMWJiIiwiZXhwIjoxNjk0NzkwNTU4LCJ1YyI6eyJzYWZlIjp0cnVlfX0.vOXtvTNZX4-H_qcHPgxMOXA3lEma3XPtK5Q36PiB1jg; strategy=local; ab.storage.sessionId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%2231ec7f48-fd10-7a63-e63e-4a55e6ad6dfb%22%2C%22e%22%3A1694705960071%2C%22c%22%3A1694704160071%2C%22l%22%3A1694704160071%7D; ab.storage.deviceId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%227cca38ec-4ce8-06d5-d010-af23569e6653%22%2C%22c%22%3A1691128725307%2C%22l%22%3A1694704160072%7D; ab.storage.userId.8d5d348c-fc26-4528-a5b4-627447ffad5a=%7B%22g%22%3A%2200bbc1a3-69e7-461b-8d05-34353f86ace0%22%2C%22c%22%3A1694528997964%2C%22l%22%3A1694704160072%7D; _ga=GA1.3.975923776.1691128716; AMP_487619ef1d=JTdCJTIyZGV2aWNlSWQlMjIlM0ElMjI5YTA2NDVkOS0yZmU5LTQzMzYtYWYyMi00M2VlMzQ3NTBlYjMlMjIlMkMlMjJ1c2VySWQlMjIlM0ElMjIwMGJiYzFhMy02OWU3LTQ2MWItOGQwNS0zNDM1M2Y4NmFjZTAlMjIlMkMlMjJzZXNzaW9uSWQlMjIlM0ExNjk0NzA0MTI1MTA3JTJDJTIyb3B0T3V0JTIyJTNBZmFsc2UlMkMlMjJsYXN0RXZlbnRUaW1lJTIyJTNBMTY5NDcwNDE2NTQ0MSUyQyUyMmxhc3RFdmVudElkJTIyJTNBMTYxJTdE; wcs_bt=s_59a6a417df3:1694704168; _ga_SRFKTMTR0R=GS1.1.1694704125.38.1.1694704169.16.0.0; _ga_5LYDPM15LW=GS1.1.1694704125.38.1.1694704169.16.0.0',
                'referer': 'https://www.kream.co.kr/products/21935',
                'sec-ch-ua': '"Chromium";v="116", "Not)A;Brand";v="24", "Google Chrome";v="116"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-origin',
                'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36',
                'x-kream-api-version': '25',
                'x-kream-client-datetime': '20230915000940+0900',
                'x-kream-device-id': 'web;75448f17-2f9f-4c5d-bb93-45a642909201',
            }
            # =======================정보가져오기========================
            try:
                basicData = GetBasicData(token,refreshToken, headers, productNo)
                with open('basicData.json', 'w', encoding='utf-8-sig') as f:
                    json.dump(basicData, f, indent=2, ensure_ascii=False)
                print("기본정보가져오기성공")
            except:
                print('기본정보가져오기실패')
                continue
            try:
                dataList1 = GetCGTransaction(daysAgo, token,refreshToken, headers, productNo)
                with open('dataList1.json', 'w', encoding='utf-8-sig') as f:
                    json.dump(dataList1, f, indent=2, ensure_ascii=False)
                print("체결정보가져오기성공")
            except:
                print("판매정보가져오기실패")
                continue
            try:
                dataList2 = GetPMTransaction(PMScroll, token,refreshToken, headers, productNo)
                with open('dataList2.json', 'w', encoding='utf-8-sig') as f:
                    json.dump(dataList2, f, indent=2, ensure_ascii=False)
                print("구매정보가져오기성공")

            except:
                print("구매정보가져오기실패")
                continue
            try:
                dataList3 = GetGMTransaction(GMScroll, token,refreshToken, headers, productNo)
                with open('dataList3.json', 'w', encoding='utf-8-sig') as f:
                    json.dump(dataList3, f, indent=2, ensure_ascii=False)
                print("구매정보가져오기성공")
            except:
                print("구매정보가져오기실패")
                continue
            time.sleep(random.randint(20,50)*0.1)

            print("데이타 저장 완료")
            # ====================정보가공하기===========================

            try:
                with open('dataList1.json', "r", encoding='utf-8-sig') as f:
                    dataList1 = json.load(f)
            except:
                dataList1=[]
            try:
                with open('dataList2.json', "r", encoding='utf-8-sig') as f:
                    dataList2 = json.load(f)
            except:
                dataList2=[]
            try:
                with open('dataList3.json', "r", encoding='utf-8-sig') as f:
                    dataList3 = json.load(f)
            except:
                dataList3=[]
            try:
                with open('basicData.json', "r", encoding='utf-8-sig') as f:
                    basicData = json.load(f)
            except:
                basicData=[]
            print("데이타 불러오기")
            # ============사이즈 같은 것들끼리 묶기
            dataListTotal = dataList1 + dataList2 + dataList3
            grouped = defaultdict(list)
            for item in dataListTotal:
                grouped[item['size']].append(item)
            sorted_grouped = sorted(grouped.items(), key=lambda x: x[0])
            results = [group[1] for group in sorted_grouped]
            print('results:',results)

            # =============각 사이즈들끼리 계산
            for index, result in enumerate(results):

                # 'category'가 'CG'인 항목 수 세기
                count = sum(1 for item in result if item['category'] == 'CG' and item['immediate'] is False)

                # 판단
                if count <= skipCount:
                    print('size:', result[0]['size'])
                    print("category가 CG/False인 항목이 3개 이하입니다.")
                    print("========================================")
                else:
                    if makeFlag == True:
                        count = 0
                        ws.title = remove_special_characters(basicData['modelCode']) + result[0]['size']

                        makeFlag = False
                    elif makeFlag == False:
                        count = 0
                        try:
                            new_sheet = wb.create_sheet(
                                remove_special_characters(basicData['modelCode']) + result[0]['size'])
                            ws = wb[remove_special_characters(basicData['modelCode']) + result[0]['size']]
                        except:
                            print('에러로넘어감')
                            continue
                        columnName = ['상품URL', '전체거래가평균(O)', '전체 거래가 평균(X)', '전체 거래가 평균(O+X)',
                                      '3개 까지의 거래가 평균(X)/전체 거래가 평균(X)',
                                      '전체 거래가 평균(O)/전체 거래가 평균(X)(=B/C)', '체결거래수(O+X)', '구매입찰 개수/판매입찰 개수',
                                      '거래가 평균(X)/매도 호가',
                                      '판매 호가(첫 행값)/발매가',
                                      '판매입찰 개수', '거래가 평균(X)/매수 호가', '구매 호가(첫 행값) / 발매가', '구매입찰개수']
                        ws.append(columnName)

                    print('size:', result[0]['size'])
                    print("category가 CG/False인 항목이 3개를 초과합니다.")
                    cg_prices = [item['price'] for item in result if item['category'] == 'CG' and item['immediate']]

                    if cg_prices:
                        valueB = int(sum(cg_prices) / len(cg_prices))
                        print(f"'category'가 'CG'이면서 'immediate'가 True인 경우의 'price'의 평균은 {valueB} 입니다.")
                    else:
                        print("해당 조건을 만족하는 항목이 없습니다.")
                        valueB = 0
                    print('valueB:', valueB)
                    cg_prices = [item['price'] for item in result if item['category'] == 'CG' and not item['immediate']]
                    # 'cg_prices' 리스트의 평균 계산
                    if cg_prices:
                        valueC = int(sum(cg_prices) / len(cg_prices))
                        print(f"'category'가 'CG'이면서 'immediate'가 False인 경우의 'price'의 평균은 {valueC} 입니다.")
                    else:
                        print("해당 조건을 만족하는 항목이 없습니다.")
                        valueC = 0
                    print('valueC:', valueC)

                    cg_prices = [item['price'] for item in result if item['category'] == 'CG']
                    # 'cg_prices' 리스트의 평균 계산
                    if cg_prices:
                        valueD = int(sum(cg_prices) / len(cg_prices))
                        print(f"'category'가 'CG'이면서 'immediate'가 False인 경우의 'price'의 평균은 {valueD} 입니다.")
                    else:
                        print("해당 조건을 만족하는 항목이 없습니다.")
                        valueD = 0
                    print('valueD:', valueD)

                    cg_prices = [item['price'] for item in result if item['category'] == 'CG' and not item['immediate']]
                    # 'cg_prices' 리스트의 평균 계산
                    cg_prices = cg_prices[:3]
                    if cg_prices:
                        valueE = int(sum(cg_prices) / len(cg_prices))
                        print(f"'category'가 'CG'이면서 'immediate'가 False인 경우의 'price'의 평균은 {valueE} 입니다.")
                    else:
                        print("해당 조건을 만족하는 항목이 없습니다.")
                        valueE = 0
                    valueE = valueE / valueC
                    print('valueE:', valueE)

                    try:
                        valueF = valueB / (valueC - 1) * 100
                    except:
                        valueF = 0
                    print('valueF:', valueF)

                    valueG = sum(1 for item in result if item['category'] == 'CG')
                    print('valueG:', valueG)

                    try:
                        lowest_price = min(item['price'] for item in result if item['category'] == 'PM')
                    except:
                        lowest_price = 0

                    try:
                        valueJ = lowest_price / basicData['originPrice']
                    except:
                        valueJ = 0
                    print("valueJ:", valueJ)

                    try:
                        valueI = valueC / lowest_price * 100
                    except:
                        valueI = 0
                    print('valueI:', valueI)

                    try:
                        valueK = sum(item['quantity'] for item in result if item['category'] == 'PM')
                    except:
                        valueK = 0
                    print('valueK:', valueK)

                    gm_items = [item for item in result if item['category'] == 'GM']

                    try:
                        if gm_items:
                            highest_price = max(gm_items, key=lambda x: x['price'])['price']
                            print(highest_price)
                        else:
                            print("No 'GM' category items found.")
                            highest_price = 0
                    except:
                        highest_price = 0
                    try:
                        valueM = highest_price / basicData['originPrice']
                    except:
                        valueM = 0
                    # print('highestPrice:',highest_price,valueB)
                    print('valueM:', valueM)

                    try:
                        valueL = valueC / highest_price * 100
                    except:
                        valueL = 0
                    print('valueL:', valueL)

                    valueN = sum(item['quantity'] for item in result if item['category'] == 'GM')

                    # 결과 출력
                    print('valueN:', valueN)

                    try:
                        valueH = valueN / valueK
                    except:
                        valueH = 0
                    print('valueH:', valueH)

                    ws.cell(row=2, column=2).value = valueB
                    ws.cell(row=2, column=3).value = valueC
                    ws.cell(row=2, column=4).value = valueD
                    ws.cell(row=2, column=5).value = valueE
                    ws.cell(row=2, column=6).value = valueF
                    ws.cell(row=2, column=7).value = valueG
                    ws.cell(row=2, column=8).value = valueH
                    ws.cell(row=2, column=9).value = valueI
                    ws.cell(row=2, column=10).value = valueJ
                    ws.cell(row=2, column=11).value = valueK
                    ws.cell(row=2, column=12).value = valueL
                    ws.cell(row=2, column=13).value = valueM
                    ws.cell(row=2, column=14).value = valueN

                    ws.cell(row=3, column=1).value = "사진"
                    ws.cell(row=3, column=2).value = "발매가"
                    ws.cell(row=3, column=3).value = "품명"
                    ws.cell(row=3, column=4).value = "모델번호"
                    ws.cell(row=3, column=5).value = "체결사이즈"
                    ws.cell(row=3, column=6).value = "거래가"
                    ws.cell(row=3, column=7).value = "번개"
                    ws.cell(row=3, column=8).value = "거래일"
                    ws.cell(row=3, column=9).value = "판매입찰사이즈"
                    ws.cell(row=3, column=10).value = "판매희망가"
                    ws.cell(row=3, column=11).value = "수량"
                    ws.cell(row=3, column=12).value = "구매입찰사이즈"
                    ws.cell(row=3, column=13).value = "구매희망가"
                    ws.cell(row=3, column=14).value = "수량"

                    ws.cell(row=4, column=1).value = basicData['imageUrl']
                    ws.cell(row=4, column=2).value = basicData['originPrice']
                    ws.cell(row=4, column=3).value = basicData['title']
                    ws.cell(row=4, column=4).value = basicData['modelCode']
                    ws.cell(row=2, column=1).value = basicData['url']

                    listCG = [[item['size'], item['price'], item['immediate'], item['transactionDate']] for item in
                              result
                              if item['category'] == 'CG']

                    for elemCG in listCG:
                        ws.cell(row=4 + count, column=5).value = elemCG[0]
                        ws.cell(row=4 + count, column=6).value = elemCG[1]
                        ws.cell(row=4 + count, column=7).value = elemCG[2]
                        ws.cell(row=4 + count, column=8).value = elemCG[3]
                        count += 1

                    listPM = [[item['size'], item['price'], item['quantity']] for item in result
                              if item['category'] == 'PM']
                    count = 0
                    for elemPM in listPM:
                        ws.cell(row=4 + count, column=9).value = elemPM[0]
                        ws.cell(row=4 + count, column=10).value = elemPM[1]
                        ws.cell(row=4 + count, column=11).value = elemPM[2]
                        count += 1

                    listGM = [[item['size'], item['price'], item['quantity']] for item in result
                              if item['category'] == 'GM']
                    count = 0
                    for elemGM in listGM:
                        ws.cell(row=4 + count, column=12).value = elemGM[0]
                        ws.cell(row=4 + count, column=13).value = elemGM[1]
                        ws.cell(row=4 + count, column=14).value = elemGM[2]
                        count += 1

                    # 첫 번째 행 고정
                    ws.freeze_panes = 'A2'

                    column_range="A:R"
                    # # 전체 열에 필터 적용
                    ws.auto_filter.ref = column_range

                    fname='result_{}.xlsx'.format(timeNow)
                    wb.save(fname)
                    print("========================================")
            time.sleep(random.randint(10,20)*0.1)

        switchD = "NONE"
        print(switchD)
        switchE = "NONE"
        print(switchE)
        switchF = "NONE"
        print(switchF)
        switchG = "NONE"
        print(switchG)
        switchH = "NONE"
        print(switchH)
        switchI = "NONE"
        print(switchI)
        switchJ = "NONE"
        print(switchJ)
        switchK = "NONE"
        print(switchK)
        switchL = "NONE"
        print(switchL)
        switchM = "NONE"
        print(switchM)
        switchN = "NONE"
        print(switchN)

        try:
            valueD =0
        except:
            valueD=0
        try:
            valueE =0
        except:
            valueE=0
        try:
            valueF = 0
        except:
            valueF=0
        try:
            valueG = 0
        except:
            valueG=0
        try:
            valueH = 0
        except:
            valueH=0
        try:
            valueI = 0
        except:
            valueI=0
        try:
            valueJ = 0
        except:
            valueJ=0
        try:
            valueK = 0
        except:
            valueK=0
        try:
            valueL = 0
        except:
            valueL=0
        try:
            valueM = 0
        except:
            valueM=0
        try:
            valueN = 0
        except:
            valueN=0


        inputValues = [
            {
                'switch': switchD,
                'name': 'valueD',
                'value': valueD},
            {
                'switch': switchE,
                'name': 'valueE',
                'value': valueE},
            {
                'switch': switchF,
                'name': 'valueF',
                'value': valueF},
            {
                'switch': switchG,
                'name': 'valueG',
                'value': valueG},
            {
                'switch': switchH,
                'name': 'valueH',
                'value': valueH},
            {
                'switch': switchI,
                'name': 'valueI',
                'value': valueI},
            {
                'switch': switchJ,
                'name': 'valueJ',
                'value': valueJ},
            {
                'switch': switchK,
                'name': 'valueK',
                'value': valueK},
            {
                'switch': switchL,
                'name': 'valueL',
                'value': valueL},
            {
                'switch': switchM,
                'name': 'valueM',
                'value': valueM},
            {
                'switch': switchN,
                'name': 'valueN',
                'value': valueN}
        ]
        searchfilename=GetSearch(fname, inputValues)
        SendMail(catIndex, searchfilename)


while True:
    timeNow=datetime.datetime.now().strftime("%H%M%S")
    targetTime="010000"
    print("현재시간:{}/목표시간:{}".format(timeNow,targetTime))
    if timeNow=="010000":
        fetchData()
    time.sleep(1)

# wb=openpyxl.load_workbook('search_result_20230915_195707.xlsx')
# ws=wb.active
# # 첫 번째 행 고정
# ws.freeze_panes = 'A2'
#
# column_range = "A:R"
# # # 전체 열에 필터 적용
# ws.auto_filter.ref = column_range
# wb.save('test111.xlsx')
